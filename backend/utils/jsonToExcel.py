import json
import openpyxl

def tecentOcrJson_to_excel(json_file, excel_file):
    # 读取JSON文件
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 解析TableDetections
    table_detections = data['TableDetections']

    # 定义一个函数来处理cells
    def process_cells_type_0(cells, start_row, start_column):
        for cell in cells:
            text = cell['Text']
            ws.cell(row=start_row, column=start_column, value=text)
            start_row += 1
        return start_row + 1, start_column

    def process_cells_type_1(cells, start_row, start_column):
        for cell in cells:
            text = cell['Text']
            col_tl = cell['ColTl']
            row_tl = cell['RowTl']
            ws.cell(row=start_row + row_tl + 1, column=col_tl + 1, value=text)
        return start_row + row_tl + 3, start_column

    # 初始化行号和列号
    row_num = 1
    column_num = 1

    # 处理header部分
    for item in table_detections:
        if item["Type"] == 0:
            row_num, column_num = process_cells_type_0(item["Cells"], row_num, column_num)
        elif item["Type"] == 1:
            row_num, column_num = process_cells_type_1(item["Cells"], row_num, column_num)

    # 保存Excel文件
    wb.save(excel_file)
    print(f"{json_file} saved to {excel_file}")

if __name__ == '__main__':
    json_file='3.json'
    excel_file = 'output3.xlsx'
    tecentOcrJson_to_excel(json_file, excel_file)